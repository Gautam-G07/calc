import openpyxl
from openpyxl import load_workbook, Workbook
wb=load_workbook(r"C:\Users\Gautam G\OneDrive\Desktop\students.xlsx")
sheet=wb.active
highest=-1
name1=""
new_wb=Workbook()
new_sheet=new_wb.active
new_sheet.append(["Name", "Marks"])
for row in sheet.iter_rows(min_row=2,values_only=True):
    name,marks=row
    
    if marks>highest:
        highest=marks
        name1=name 
    if marks>=75:
        new_sheet.append([name, marks])  
print(f"The highest marks are {highest} obtained by {name1}.")
new_wb.save(r"C:\Users\Gautam G\OneDrive\Desktop\top_students.xlsx")
print("New Excel file with top students created successfully.")